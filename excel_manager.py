import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ARCHIVO = "pokedex_datos.xlsx"

def inicializar_excel():
    if not os.path.exists(ARCHIVO):
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        ws.append(["ID", "Nombre (Link)", "Tipos", "HP", "Poder Total", "Fecha", "Estado"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(ARCHIVO)

def guardar_pokemon_excel(pokemon, estado="Registrado"):
    wb = load_workbook(ARCHIVO)
    ws = wb.active
    poder_total = sum(pokemon.stats.values())
    ws.append([
        pokemon.id, 
        pokemon.nombre, 
        pokemon.tipos, 
        pokemon.hp_max, 
        poder_total, 
        datetime.now().strftime("%d/%m/%Y %H:%M"), 
        estado
    ])
    celda_nombre = ws.cell(row=ws.max_row, column=2)
    ruta_img = f"imagenes/{pokemon.nombre}.png"
    if os.path.exists(ruta_img):
        celda_nombre.hyperlink = ruta_img
        celda_nombre.font = Font(color="0000FF", underline="single")
    wb.save(ARCHIVO)

def registrar_victoria_excel(nombre_ganador):
    if not os.path.exists(ARCHIVO): return
    wb = load_workbook(ARCHIVO)
    ws = wb.active
    for row in reversed(list(ws.iter_rows(min_row=2))):
        if row[1].value == nombre_ganador:
            row[6].value = "GANADOR"
            row[6].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            break
    wb.save(ARCHIVO)

def obtener_resumen_pokedex():
    if not os.path.exists(ARCHIVO): return {}
    wb = load_workbook(ARCHIVO)
    ws = wb.active
    resumen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            tipo = row[2].split(",")[0].strip()
            resumen[tipo] = resumen.get(tipo, 0) + 1
    return resumen